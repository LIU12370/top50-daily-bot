#!/usr/bin/env python3
"""
TOP50 Hot Articles Web Application
Flask backend: scraping, scoring, Excel generation
"""

import os, re, json, time, random, threading
from datetime import datetime, timedelta
from io import BytesIO
from flask import Flask, jsonify, send_file, render_template, request
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder="static", template_folder="static")

# ── Cache ──────────────────────────────────────────────────
_cache = {"data": None, "time": None, "excel": None}
CACHE_TTL = 0  # 每次点击都生成最新结果，不使用缓存

# ── 五大领域定义 ──────────────────────────────────────────
DOMAINS = {
    "政治": {
        "label": "政治",
        "color": "#B8432F",
        # 语义扩展词：用于将微博热搜归类到该领域
        "classify_keywords": [
            "政治", "政策", "两会", "总书记", "国务院", "中央", "党",
            "外交", "峰会", "制裁", "谈判", "条约", "协议", "领导人",
            "选举", "议会", "国会", "白宫", "总统", "首相", "普京",
            "特朗普", "拜登", "欧盟", "联合国", "北约", "G7", "G20",
            "台海", "南海", "朝鲜", "半岛", "中美", "中欧", "中俄",
            "一带一路", "反腐", "纪委", "巡视", "法治", "立法",
            "人大", "政协", "改革", "统战", "港澳", "治理",
        ],
    },
    "财经": {
        "label": "财经",
        "color": "#C8A951",
        "classify_keywords": [
            "财经", "经济", "金融", "股市", "A股", "港股", "美股",
            "基金", "债券", "利率", "降息", "加息", "央行", "美联储",
            "LPR", "GDP", "CPI", "通胀", "通缩", "汇率", "美元", "人民币",
            "黄金", "金价", "原油", "油价", "期货", "大宗商品",
            "楼市", "房价", "地产", "IPO", "上市", "融资", "并购",
            "财报", "营收", "利润", "市值", "估值", "PE",
            "银行", "保险", "证券", "信托", "理财", "投资",
            "摩根", "高盛", "瑞银", "花旗", "巴菲特",
            "光伏", "锂电", "储能", "新能源", "碳中和",
        ],
    },
    "AI": {
        "label": "AI",
        "color": "#6C5CE7",
        "classify_keywords": [
            "AI", "人工智能", "大模型", "GPT", "Claude", "Gemini",
            "OpenAI", "Anthropic", "DeepSeek", "通义", "千问", "文心",
            "Cursor", "Copilot", "Agent", "智能体", "龙虾", "OpenClaw",
            "机器学习", "深度学习", "强化学习", "神经网络", "Transformer",
            "AIGC", "生成式", "Sora", "视频生成", "图像生成",
            "token", "推理", "训练", "微调", "开源模型",
            "MiMo", "Qwen", "Llama", "Mistral",
            "氛围编程", "Vibe Coding", "AI编程",
            "脑机接口", "AGI", "超级智能",
        ],
    },
    "科技": {
        "label": "科技",
        "color": "#2A7B6E",
        "classify_keywords": [
            "科技", "技术", "互联网", "数字化", "云计算", "大数据",
            "芯片", "半导体", "GPU", "英伟达", "NVIDIA", "AMD", "台积电",
            "苹果", "iPhone", "华为", "小米", "OPPO", "vivo", "三星",
            "腾讯", "阿里", "字节", "百度", "美团", "京东", "拼多多",
            "马斯克", "特斯拉", "SpaceX", "星链",
            "自动驾驶", "无人驾驶", "新能源车", "电动车",
            "机器人", "具身智能", "无人机", "大疆",
            "5G", "6G", "量子计算", "区块链", "元宇宙", "VR", "AR",
            "SaaS", "开源", "GitHub", "飞书", "钉钉",
            "生物科技", "基因", "mRNA", "航天", "卫星",
        ],
    },
    "军事": {
        "label": "军事",
        "color": "#4A5568",
        "classify_keywords": [
            "军事", "国防", "军队", "解放军", "海军", "空军",
            "导弹", "核武", "航母", "战斗机", "歼", "轰",
            "战争", "冲突", "空袭", "军演", "演习",
            "伊朗", "以色列", "中东", "巴以", "哈马斯", "真主党",
            "俄乌", "乌克兰", "俄罗斯", "北约", "军援",
            "霍尔木兹", "红海", "胡塞", "叙利亚", "阿富汗",
            "武器", "军工", "军费", "征兵", "兵役",
            "核潜艇", "隐身", "高超音速", "防空", "雷达",
            "网络战", "太空军", "军事卫星",
        ],
    },
}

# 所有领域名称列表
DOMAIN_NAMES = list(DOMAINS.keys())

# ── Target WeChat Public Accounts ──────────────────────────
TARGET_ACCOUNTS = [
    "三联生活周刊", "视觉志", "网易上流", "闻旅", "新周刊",
    "三联生活实验室", "正解局", "混知", "九行", "地道风物",
    "刺猬公社", "BT财经", "华尔街见闻", "腾讯科技", "砺石商业评论",
    "功夫财经", "光子星球", "新智元", "量子位", "哈佛商业评论",
    "外滩TheBund", "新世界相"
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}

# ── WeChat MP Backend API Configuration ────────────────────
# 从环境变量读取登录凭证（在 Render 的 Environment 中配置）
# WECHAT_COOKIES: 登录 mp.weixin.qq.com 后浏览器的完整 Cookie 字符串
# WECHAT_TOKEN:   登录后 URL 中的 token 参数（纯数字）
WECHAT_COOKIES = os.environ.get("WECHAT_COOKIES", "")
WECHAT_TOKEN = os.environ.get("WECHAT_TOKEN", "")

# fakeid 缓存：已知的 fakeid 硬编码 + 运行时自动查询的缓存
# 硬编码的会跳过 searchbiz 查询，节省请求次数
FAKEID_MAP = {
    "刺猬公社": "MzkxNzAwMDkwNQ==",
    "BT财经": "MzA5OTg2NzA5Ng==",
    "华尔街见闻": "MjM5NzAwMzU0MA==",
    "腾讯科技": "Mjc1NjM3MjY2MA==",
    "功夫财经": "MzIzOTA3NTA5Mg==",
    "光子星球": "MzA4MjUxODMwMg==",
    "新智元": "MzI3MTA0MTk1MA==",
    "量子位": "MzIzNjc1NzUzMw==",
    "砺石商业评论": "MzIyMDMyNTMwMw==",
}
# 环境变量补充（JSON 格式，会合并到 FAKEID_MAP）
_fakeid_env = os.environ.get("WECHAT_FAKEIDS", "{}")
try:
    FAKEID_MAP.update(json.loads(_fakeid_env))
except (json.JSONDecodeError, TypeError):
    pass


def _classify_keyword(keyword):
    """将一个热搜关键词归类到五大领域之一，返回领域名称"""
    kw_lower = keyword.lower()
    scores = {}
    for domain, cfg in DOMAINS.items():
        score = 0
        for ck in cfg["classify_keywords"]:
            if ck.lower() in kw_lower or kw_lower in ck.lower():
                score += 1
        scores[domain] = score
    best = max(scores, key=scores.get)
    if scores[best] > 0:
        return best
    # 无法分类的默认归入"科技"
    return "科技"


def scrape_weibo_hot():
    """
    抓取微博实时热搜，分类到五大领域。
    返回: {"政治": [...], "财经": [...], "AI": [...], "科技": [...], "军事": [...]}
    每个领域的关键词列表，语义扩展后用于文章匹配。
    """
    raw_keywords = []
    try:
        url = "https://weibo.com/ajax/side/hotSearch"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        data = resp.json()
        for item in data.get("data", {}).get("realtime", []):
            word = item.get("word", "").strip()
            if word and len(word) >= 2:
                raw_keywords.append(word)
        print(f"[Weibo API] fetched {len(raw_keywords)} live hot keywords")
    except Exception as e:
        print(f"[Weibo API] error: {e}")

    # 热搜词补充：热门话题 API
    try:
        url2 = "https://weibo.com/ajax/statuses/hot_band"
        resp2 = requests.get(url2, headers=HEADERS, timeout=10)
        data2 = resp2.json()
        seen = set(k.lower() for k in raw_keywords)
        for item in data2.get("data", {}).get("band_list", []):
            word = item.get("word", "").strip()
            if word and len(word) >= 2 and word.lower() not in seen:
                raw_keywords.append(word)
                seen.add(word.lower())
    except Exception:
        pass

    # 分类到五大领域
    domain_keywords = {d: [] for d in DOMAIN_NAMES}
    for kw in raw_keywords:
        domain = _classify_keyword(kw)
        domain_keywords[domain].append(kw)

    # 确保每个领域至少有基础关键词（语义扩展兜底）
    domain_fallback = {
        "政治": ["政策", "外交", "两会", "改革", "治理", "中美关系"],
        "财经": ["股市", "经济", "金融", "投资", "央行", "汇率"],
        "AI":   ["AI", "人工智能", "大模型", "智能体", "Agent"],
        "科技": ["科技", "芯片", "互联网", "创新", "数字化"],
        "军事": ["军事", "国防", "冲突", "安全", "战争"],
    }
    for domain in DOMAIN_NAMES:
        if len(domain_keywords[domain]) < 3:
            seen = set(k.lower() for k in domain_keywords[domain])
            for fb in domain_fallback.get(domain, []):
                if fb.lower() not in seen:
                    domain_keywords[domain].append(fb)
                    seen.add(fb.lower())

    for d in DOMAIN_NAMES:
        print(f"[Keywords] {d}: {len(domain_keywords[d])} keywords")

    return domain_keywords


# ── WeChat MP Backend API Scraper ─────────────────────────

def _mp_headers():
    """Build request headers for MP backend API"""
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 "
                       "MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI",
        "Referer": f"https://mp.weixin.qq.com/cgi-bin/appmsg?"
                    f"t=media/appmsg_edit_v2&action=edit&token={WECHAT_TOKEN}",
        "Cookie": WECHAT_COOKIES,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
    }


def _check_mp_ret(data, context=""):
    """
    检查 MP API 返回的错误码。
    返回: "ok" | "rate_limited" | "expired" | "error"
    """
    ret = data.get("base_resp", {}).get("ret", 0)
    if ret == 0:
        return "ok"
    if ret == 200013:
        print(f"[MP API] RATE LIMITED (200013) {context}. Stopping.")
        return "rate_limited"
    if ret == 200040:
        print(f"[MP API] COOKIE/TOKEN EXPIRED (200040). Please refresh credentials.")
        return "expired"
    print(f"[MP API] error ret={ret} {context}")
    return "error"


def _get_fakeid(account_name, headers):
    """
    通过 searchbiz 接口自动查询公众号的 fakeid。
    严格匹配 nickname == account_name，防止重名误匹配。
    返回 fakeid 字符串，失败返回 None。
    """
    search_url = "https://mp.weixin.qq.com/cgi-bin/searchbiz"
    params = {
        "action": "search_biz",
        "token": WECHAT_TOKEN,
        "lang": "zh_CN",
        "f": "json",
        "ajax": "1",
        "query": account_name,
        "begin": "0",
        "count": "5",
    }

    try:
        resp = requests.get(search_url, params=params, headers=headers, timeout=15)
        data = resp.json()

        status = _check_mp_ret(data, f"searchbiz '{account_name}'")
        if status != "ok":
            return None if status == "error" else status  # 传播 rate_limited/expired

        biz_list = data.get("list", [])
        if not biz_list:
            print(f"[searchbiz] '{account_name}': no results")
            return None

        # 严格匹配：nickname 必须完全一致
        for item in biz_list:
            nickname = item.get("nickname", "").strip()
            if nickname == account_name:
                fid = item.get("fakeid", "")
                print(f"[searchbiz] '{account_name}' -> fakeid={fid}")
                return fid

        # 宽松匹配：处理 "外滩TheBund" 等中英混合名称
        for item in biz_list:
            nickname = item.get("nickname", "").strip()
            # 互相包含即视为匹配
            if account_name in nickname or nickname in account_name:
                fid = item.get("fakeid", "")
                print(f"[searchbiz] '{account_name}' ~ '{nickname}' -> fakeid={fid}")
                return fid

        print(f"[searchbiz] '{account_name}': no exact match in {[i.get('nickname') for i in biz_list]}")
        return None

    except requests.exceptions.RequestException as e:
        print(f"[searchbiz] network error '{account_name}': {e}")
        return None
    except (json.JSONDecodeError, KeyError) as e:
        print(f"[searchbiz] parse error '{account_name}': {e}")
        return None


def _fetch_articles_for_fakeid(account, fakeid, headers, cutoff):
    """
    用 fakeid 抓取该公众号最近10篇文章（2页 x 5条）。
    返回 articles 列表。
    如果遇到 rate_limited/expired 返回对应字符串信号。
    """
    articles = []
    now = datetime.now()
    api_url = "https://mp.weixin.qq.com/cgi-bin/appmsg"

    for page in range(2):
        params = {
            "action": "list_ex",
            "begin": page * 5,
            "count": 5,
            "fakeid": fakeid,
            "type": "9",
            "query": "",
            "token": WECHAT_TOKEN,
            "lang": "zh_CN",
            "f": "json",
            "ajax": "1",
        }

        try:
            resp = requests.get(api_url, params=params, headers=headers, timeout=15)
            data = resp.json()

            status = _check_mp_ret(data, f"appmsg '{account}' page {page}")
            if status == "rate_limited":
                return "rate_limited"
            if status == "expired":
                return "expired"
            if status == "error":
                break

            app_msg_list = data.get("app_msg_list", [])
            if not app_msg_list:
                break

            for item in app_msg_list:
                title = item.get("title", "").strip()
                digest = item.get("digest", "")
                create_time = item.get("create_time", 0)

                if not title or len(title) < 5:
                    continue

                pub_dt = datetime.fromtimestamp(create_time) if create_time else now

                # 严格72小时过滤
                if pub_dt < cutoff:
                    continue

                articles.append({
                    "account": account,
                    "title": title,
                    "summary": digest[:200] if digest else "",
                    "pub_time": pub_dt.strftime("%Y-%m-%d %H:%M"),
                    "_pub_dt": pub_dt,
                    "_source": "mp_api",
                })

            print(f"[MP API] {account} page {page}: {len(app_msg_list)} items, {len(articles)} within 72h")

        except requests.exceptions.RequestException as e:
            print(f"[MP API] network error '{account}': {e}")
            break
        except (json.JSONDecodeError, KeyError) as e:
            print(f"[MP API] parse error '{account}': {e}")
            break

        # 每页间隔 2-4 秒
        time.sleep(random.uniform(2, 4))

    return articles


def _scrape_mp_api():
    """
    完整 MP 后台抓取流程：
    1. 优先处理有 fakeid 的账号（快速）
    2. 没有 fakeid 的通过 searchbiz 查询（较慢，有风控）
    3. 账号间间隔 3-5 秒（已有 fakeid）或 8-12 秒（需 searchbiz）
    """
    if not WECHAT_COOKIES or not WECHAT_TOKEN:
        print("[MP API] skipped: WECHAT_COOKIES or WECHAT_TOKEN not configured")
        return []

    articles = []
    now = datetime.now()
    cutoff = now - timedelta(hours=72)
    headers = _mp_headers()
    searched_count = 0

    # 先处理有 fakeid 的账号（速度快，间隔短）
    accounts_with_fid = [a for a in TARGET_ACCOUNTS if a in FAKEID_MAP]
    accounts_without_fid = [a for a in TARGET_ACCOUNTS if a not in FAKEID_MAP]

    for idx, account in enumerate(accounts_with_fid):
        fakeid = FAKEID_MAP[account]
        _task["progress"] = f"正在抓取 {account}...（{idx+1}/{len(accounts_with_fid)}）"
        print(f"[MP API] {account}: using cached fakeid")

        result = _fetch_articles_for_fakeid(account, fakeid, headers, cutoff)

        if result == "rate_limited":
            print("[MP API] appmsg rate limited, returning collected data")
            return articles
        if result == "expired":
            print("[MP API] credentials expired, returning collected data")
            return articles

        if isinstance(result, list):
            articles.extend(result)

        # 已有 fakeid 的账号间隔短一些
        sleep_time = random.uniform(3, 5)
        print(f"[MP API] sleeping {sleep_time:.0f}s before next account...")
        time.sleep(sleep_time)

    # 再处理需要 searchbiz 的账号
    for account in accounts_without_fid:
        print(f"[MP API] {account}: searching fakeid via searchbiz...")
        result = _get_fakeid(account, headers)

        if result == "rate_limited":
            print("[MP API] searchbiz rate limited, returning collected data")
            return articles
        if result == "expired":
            print("[MP API] credentials expired, returning collected data")
            return articles
        if not result:
            print(f"[MP API] {account}: fakeid not found, skipping")
            time.sleep(random.uniform(5, 8))
            continue

        fakeid = result
        FAKEID_MAP[account] = fakeid
        searched_count += 1

        # searchbiz 后等一下再抓文章
        time.sleep(random.uniform(5, 8))

        result = _fetch_articles_for_fakeid(account, fakeid, headers, cutoff)

        if result == "rate_limited":
            print("[MP API] appmsg rate limited, returning collected data")
            return articles
        if result == "expired":
            print("[MP API] credentials expired, returning collected data")
            return articles

        if isinstance(result, list):
            articles.extend(result)

        sleep_time = random.uniform(8, 12)
        print(f"[MP API] sleeping {sleep_time:.0f}s before next account...")
        time.sleep(sleep_time)

    print(f"[MP API] done: {len(articles)} articles | "
          f"{len(FAKEID_MAP)} fakeids cached | "
          f"{searched_count} new lookups this run")
    return articles


def _parse_relative_time(text):
    """Parse Chinese relative time strings into datetime"""
    now = datetime.now()
    text = text.strip()
    m = re.search(r'(\d+)\s*分钟前', text)
    if m:
        return now - timedelta(minutes=int(m.group(1)))
    m = re.search(r'(\d+)\s*小时前', text)
    if m:
        return now - timedelta(hours=int(m.group(1)))
    m = re.search(r'(\d+)\s*天前', text)
    if m:
        return now - timedelta(days=int(m.group(1)))
    m = re.search(r'今天\s*(\d{1,2}):(\d{2})', text)
    if m:
        return now.replace(hour=int(m.group(1)), minute=int(m.group(2)), second=0)
    m = re.search(r'昨天', text)
    if m:
        return now - timedelta(days=1)
    # Try parsing absolute dates
    for fmt in ["%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y年%m月%d日", "%m月%d日", "%m-%d"]:
        try:
            parsed = datetime.strptime(text, fmt)
            if parsed.year == 1900:
                parsed = parsed.replace(year=now.year)
            return parsed
        except ValueError:
            continue
    # Extract any date-like pattern
    m = re.search(r'(\d{4})[-.年](\d{1,2})[-.月](\d{1,2})', text)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    m = re.search(r'(\d{1,2})[-.月](\d{1,2})', text)
    if m:
        try:
            return datetime(now.year, int(m.group(1)), int(m.group(2)))
        except ValueError:
            pass
    return None


def _is_within_days(pub_dt, days=3):
    """Check if a datetime is within the last N days"""
    if not pub_dt:
        return False
    cutoff = datetime.now() - timedelta(days=days)
    return pub_dt >= cutoff


def search_articles_for_account(account):
    """Search recent articles for a specific public account via Sogou"""
    articles = []
    try:
        # Sogou WeChat article search
        query = f'{account} {datetime.now().strftime("%Y年%m月")}'
        search_url = "https://weixin.sogou.com/weixin"
        params = {"type": "2", "query": query, "ie": "utf8"}
        resp = requests.get(search_url, params=params, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, "lxml")

        for item in soup.select(".news-list li, .news-box li, .txt-box, .vrwrap, .rb"):
            title_el = item.select_one("h3 a, .tit a, .vr-title a, h4 a")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            # Clean up title artifacts
            title = re.sub(r'\s+', ' ', title).strip()

            summary_el = item.select_one("p.txt-info, .txt-info, .str_info, .space-txt, p")
            summary = summary_el.get_text(strip=True) if summary_el else ""
            summary = summary[:200]

            # Extract time info
            time_el = item.select_one(".s-p, .s2, .news-when, span.time, .time-box")
            time_text = time_el.get_text(strip=True) if time_el else ""
            pub_dt = _parse_relative_time(time_text) if time_text else None

            # Account name verification
            account_el = item.select_one(".account, .all-txt, .s-p a")
            source_text = account_el.get_text(strip=True) if account_el else ""

            check_name = account.replace("TheBund", "")
            if len(title) > 5 and (check_name in (title + summary + source_text)):
                pub_time_str = pub_dt.strftime("%Y-%m-%d %H:%M") if pub_dt else datetime.now().strftime("%Y-%m-%d %H:%M")
                articles.append({
                    "account": account,
                    "title": title,
                    "summary": summary,
                    "pub_time": pub_time_str,
                    "_pub_dt": pub_dt,
                })
    except Exception as e:
        print(f"[Sogou {account}] {e}")

    # Also try general web search as backup
    if len(articles) < 2:
        try:
            query2 = f'"{account}" 微信公众号 最新文章'
            resp2 = requests.get("https://www.sogou.com/web", params={"query": query2}, headers=HEADERS, timeout=8)
            soup2 = BeautifulSoup(resp2.text, "lxml")
            for item in soup2.select(".vrwrap, .rb"):
                title_el = item.select_one("h3 a, .vr-title a")
                if not title_el:
                    continue
                title = title_el.get_text(strip=True)
                title = re.sub(r'\s+', ' ', title).strip()
                summary_el = item.select_one(".str_info, .space-txt, .str-text, p")
                summary = summary_el.get_text(strip=True)[:200] if summary_el else ""

                check_name = account.replace("TheBund", "")
                if len(title) > 5 and check_name in (title + summary):
                    # Deduplicate
                    if not any(a["title"] == title for a in articles):
                        articles.append({
                            "account": account,
                            "title": title,
                            "summary": summary,
                            "pub_time": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "_pub_dt": None,
                        })
        except Exception as e:
            print(f"[Web search {account}] {e}")

    return articles


def _scrape_tophub():
    """Scrape tophub.today for trending WeChat articles"""
    articles = []
    target_set = set(TARGET_ACCOUNTS)
    target_set_lower = set(a.lower() for a in TARGET_ACCOUNTS)

    # tophub WeChat section
    urls = [
        "https://tophub.today/n/WnBe01o371",   # WeChat
        "https://tophub.today/n/Y2KeDGQdNP",   # 36kr
    ]
    for tophub_url in urls:
        try:
            resp = requests.get(tophub_url, headers=HEADERS, timeout=10)
            soup = BeautifulSoup(resp.text, "lxml")
            for item in soup.select("table tbody tr, .Zd-p-Sc .cc-dc"):
                tds = item.select("td")
                if len(tds) >= 2:
                    title = tds[1].get_text(strip=True) if len(tds) > 1 else ""
                    # Some tophub layouts have link text
                    link_el = item.select_one("a")
                    if link_el:
                        title = link_el.get_text(strip=True) or title
                    time_el = item.select_one("td:last-child")
                    time_text = time_el.get_text(strip=True) if time_el else ""

                    # Try to match to target accounts
                    for acc in TARGET_ACCOUNTS:
                        check = acc.replace("TheBund", "")
                        if check in title and len(title) > 5:
                            pub_dt = _parse_relative_time(time_text)
                            articles.append({
                                "account": acc,
                                "title": title,
                                "summary": "",
                                "pub_time": pub_dt.strftime("%Y-%m-%d %H:%M") if pub_dt else datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "_pub_dt": pub_dt,
                            })
                            break
        except Exception as e:
            print(f"[TopHub] {e}")

    return articles


def _scrape_aggregator_sites():
    """Scrape 36kr and huxiu for republished articles from target accounts"""
    articles = []
    now = datetime.now()

    # 36kr recent articles
    try:
        resp = requests.get("https://36kr.com/information/web_news/", headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, "lxml")
        for item in soup.select(".article-item, .information-flow-item, .kr-flow-bar-item"):
            title_el = item.select_one("a.article-item-title, .article-item-title, a")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            summary_el = item.select_one(".article-item-description, .summary")
            summary = summary_el.get_text(strip=True)[:200] if summary_el else ""
            time_el = item.select_one(".kr-flow-bar-time, time, .time")
            time_text = time_el.get_text(strip=True) if time_el else ""
            pub_dt = _parse_relative_time(time_text)

            # Try to match to target accounts
            for acc in TARGET_ACCOUNTS:
                check = acc.replace("TheBund", "")
                if check in (title + summary):
                    articles.append({
                        "account": acc,
                        "title": title,
                        "summary": summary,
                        "pub_time": pub_dt.strftime("%Y-%m-%d %H:%M") if pub_dt else now.strftime("%Y-%m-%d %H:%M"),
                        "_pub_dt": pub_dt,
                    })
                    break
            else:
                # Even if not matched to target account, keep tech/finance articles
                if any(kw in title for kw in ["AI", "融资", "科技", "芯片", "模型", "机器人"]):
                    articles.append({
                        "account": "36氪",
                        "title": title,
                        "summary": summary,
                        "pub_time": pub_dt.strftime("%Y-%m-%d %H:%M") if pub_dt else now.strftime("%Y-%m-%d %H:%M"),
                        "_pub_dt": pub_dt,
                    })
    except Exception as e:
        print(f"[36kr] {e}")

    # Huxiu recent articles
    try:
        resp = requests.get("https://www.huxiu.com/", headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, "lxml")
        for item in soup.select(".recommend-article-item, .article-item, .nfmore-item"):
            title_el = item.select_one("a, h2 a, .title")
            if not title_el:
                continue
            title = title_el.get_text(strip=True)
            summary_el = item.select_one(".article-summary, .summary, .desc")
            summary = summary_el.get_text(strip=True)[:200] if summary_el else ""
            time_el = item.select_one("time, .time, .article-time")
            time_text = time_el.get_text(strip=True) if time_el else ""
            pub_dt = _parse_relative_time(time_text)

            for acc in TARGET_ACCOUNTS:
                check = acc.replace("TheBund", "")
                if check in (title + summary):
                    articles.append({
                        "account": acc,
                        "title": title,
                        "summary": summary,
                        "pub_time": pub_dt.strftime("%Y-%m-%d %H:%M") if pub_dt else now.strftime("%Y-%m-%d %H:%M"),
                        "_pub_dt": pub_dt,
                    })
                    break
    except Exception as e:
        print(f"[Huxiu] {e}")

    return articles


def _scrape_known_sites():
    """Scrape accounts that have their own public websites"""
    articles = []
    now = datetime.now()

    # Map of accounts to their public website article list URLs
    site_map = {
        "量子位": "https://www.qbitai.com/",
        "华尔街见闻": "https://wallstreetcn.com/news/global",
        "新智元": "https://www.ainewera.com/",
        "刺猬公社": "https://www.ciweigongshe.net/",
        "36氪": "https://36kr.com/",
    }

    for account, site_url in site_map.items():
        if account not in TARGET_ACCOUNTS and account != "36氪":
            continue
        try:
            resp = requests.get(site_url, headers=HEADERS, timeout=10)
            soup = BeautifulSoup(resp.text, "lxml")

            # Generic article extraction
            for item in soup.select("article, .article-item, .news-item, .post-item, .list-item, .feed-item"):
                title_el = item.select_one("h2 a, h3 a, .title a, a.title, h2, h3")
                if not title_el:
                    continue
                title = title_el.get_text(strip=True)
                if len(title) < 6:
                    continue
                title = re.sub(r'\s+', ' ', title).strip()

                summary_el = item.select_one("p, .summary, .desc, .content-text")
                summary = summary_el.get_text(strip=True)[:200] if summary_el else ""

                time_el = item.select_one("time, .time, .date, .meta-time, span[data-time]")
                time_text = time_el.get_text(strip=True) if time_el else ""
                pub_dt = _parse_relative_time(time_text)

                articles.append({
                    "account": account,
                    "title": title,
                    "summary": summary,
                    "pub_time": pub_dt.strftime("%Y-%m-%d %H:%M") if pub_dt else now.strftime("%Y-%m-%d %H:%M"),
                    "_pub_dt": pub_dt,
                })

            time.sleep(0.3)
        except Exception as e:
            print(f"[Site {account}] {e}")

    return articles


def search_articles_bulk():
    """Search articles from multiple real sources, strictly filtered to 48 hours"""
    all_articles = []
    existing_titles = set()
    now = datetime.now()
    cutoff = now - timedelta(hours=72)

    def add_unique(articles_list):
        for art in articles_list:
            title = art["title"].strip()
            if title in existing_titles:
                continue
            if len(title) < 6:
                continue
            existing_titles.add(title)
            all_articles.append(art)

    # Source 0 (PRIMARY): WeChat MP Backend API — real article data
    print("[Pipeline] Trying MP backend API...")
    try:
        mp_results = _scrape_mp_api()
        add_unique(mp_results)
    except Exception as e:
        print(f"[Pipeline] MP API error: {e}")
    print(f"[Pipeline] After MP API: {len(all_articles)} articles")

    # Source 1: Sogou WeChat search for each target account
    print("[Pipeline] Searching Sogou for target accounts...")
    for acc in TARGET_ACCOUNTS:
        try:
            results = search_articles_for_account(acc)
            add_unique(results)
            time.sleep(0.5)
        except Exception:
            pass

    print(f"[Pipeline] After Sogou: {len(all_articles)} articles")

    # Source 2: TopHub aggregator
    print("[Pipeline] Scraping TopHub...")
    try:
        tophub_results = _scrape_tophub()
        add_unique(tophub_results)
    except Exception:
        pass
    print(f"[Pipeline] After TopHub: {len(all_articles)} articles")

    # Source 3: 36kr and Huxiu
    print("[Pipeline] Scraping aggregator sites...")
    try:
        agg_results = _scrape_aggregator_sites()
        add_unique(agg_results)
    except Exception:
        pass
    print(f"[Pipeline] After aggregators: {len(all_articles)} articles")

    # Source 4: Known public websites
    print("[Pipeline] Scraping known sites...")
    try:
        site_results = _scrape_known_sites()
        add_unique(site_results)
    except Exception:
        pass
    print(f"[Pipeline] After known sites: {len(all_articles)} articles")

    # STRICT 72-hour freshness filter
    filtered = []
    for art in all_articles:
        pub_dt = art.get("_pub_dt")
        # 没有可验证的发布时间 → 丢弃（防止过期文章混入）
        if not pub_dt:
            continue
        if pub_dt < cutoff:
            continue  # Discard articles older than 72 hours
        filtered.append(art)
        # Remove internal field
        art.pop("_pub_dt", None)

    print(f"[Pipeline] After 72h filter: {len(filtered)} articles (from {len(all_articles)})")
    return filtered


def match_keywords_with_domain(article, domain_keywords):
    """
    匹配文章与各领域热搜关键词。
    返回: (matched_list, primary_domain, primary_keyword)
      - matched_list: [{"keyword": "xxx", "domain": "AI"}, ...]
      - primary_domain: 匹配最多的领域名称
      - primary_keyword: 第一个匹配到的热搜词
    """
    text = (article["title"] + " " + article.get("summary", "")).lower()
    matched_list = []
    domain_counts = {d: 0 for d in DOMAIN_NAMES}

    for domain, keywords in domain_keywords.items():
        for kw in keywords:
            if kw.lower() in text:
                matched_list.append({"keyword": kw, "domain": domain})
                domain_counts[domain] += 1

    # 用领域分类扩展词做二次匹配（语义扩展）
    for domain, cfg in DOMAINS.items():
        for ck in cfg["classify_keywords"]:
            if ck.lower() in text and not any(m["keyword"].lower() == ck.lower() for m in matched_list):
                matched_list.append({"keyword": ck, "domain": domain})
                domain_counts[domain] += 1

    primary_domain = max(domain_counts, key=domain_counts.get) if any(domain_counts.values()) else "科技"
    primary_keyword = matched_list[0]["keyword"] if matched_list else ""

    return matched_list, primary_domain, primary_keyword


def verify_article_source(article):
    """
    验证文章是否真实来源于指定公众号。
    返回 True 表示通过验证。
    """
    account = article.get("account", "")
    title = article.get("title", "")

    # MP API 来源的文章天然可信（通过 fakeid 精确查询）
    if article.get("_source") == "mp_api":
        return True

    # 必须属于 TARGET_ACCOUNTS
    if account not in TARGET_ACCOUNTS:
        return False

    # 标题不能过短
    if len(title) < 6:
        return False

    # 排除明显的广告/推广内容
    spam_patterns = ["广告", "推广", "赞助", "合作推荐", "点击领取", "限时免费"]
    if any(sp in title for sp in spam_patterns):
        return False

    # 排除每日资讯合集/汇总类文章（低原创价值）
    roundup_patterns = [
        "每日资讯", "日报", "周报", "早报", "晚报", "资讯合集",
        "一周回顾", "本周盘点", "今日头条汇总", "热点速递",
        "AI资讯", "科技资讯", "财经资讯", "每日精选",
        "Top 榜", "TOP榜", "雷达 Top", "资讯雷达",
        "一文速览", "快讯合集", "要闻速递", "每周精选",
        "Morning Brief", "Daily Brief", "Weekly Digest",
    ]
    if any(rp in title for rp in roundup_patterns):
        return False

    return True


def ai_score(article):
    score = 5.0
    title = article["title"]
    summary = article["summary"]
    combined = title + " " + summary

    if len(summary) > 60: score += 0.5
    if len(summary) > 100: score += 0.5
    if re.search(r'\d+', combined): score += 0.5

    opinion_words = ["深度", "分析", "揭秘", "逻辑", "背后", "颠覆", "反思",
                     "警告", "危机", "革命", "突破", "创新", "重磅", "首次",
                     "解析", "判断", "预测", "推演", "变局", "登顶"]
    score += min(sum(1 for w in opinion_words if w in combined) * 0.3, 1.5)

    if any(c in title for c in "？！"): score += 0.3
    if any(c in title for c in "：｜"): score += 0.2
    if 10 <= len(title) <= 30: score += 0.3

    hot_topics = ["AI", "人工智能", "大模型", "芯片", "光伏", "特斯拉", "马斯克",
                  "伊朗", "原油", "黄金", "A股", "腾讯", "阿里", "英伟达",
                  "OpenClaw", "龙虾", "Agent", "机器人", "具身智能"]
    score += min(sum(1 for t in hot_topics if t.lower() in combined.lower()) * 0.2, 1.0)

    authority = {"华尔街见闻": 0.5, "三联生活周刊": 0.5, "量子位": 0.4,
                 "新智元": 0.3, "腾讯科技": 0.4, "哈佛商业评论": 0.5,
                 "正解局": 0.3, "光子星球": 0.3, "BT财经": 0.2,
                 "刺猬公社": 0.3, "功夫财经": 0.2}
    score += authority.get(article["account"], 0.1)

    return round(min(max(score, 1.0), 10.0), 1)


def time_weight(pub_time_str):
    try:
        pub = datetime.strptime(pub_time_str, "%Y-%m-%d %H:%M")
    except:
        pub = datetime.strptime(pub_time_str.split(" ")[0], "%Y-%m-%d")
    hours = (datetime.now() - pub).total_seconds() / 3600
    if hours <= 6: return 3.0
    elif hours <= 12: return 2.5
    elif hours <= 24: return 2.0
    elif hours <= 48: return 1.5
    elif hours <= 72: return 1.0
    else: return 0.5


def generate_top50():
    """Full pipeline: scrape -> match -> score -> verify -> balanced rank"""
    # Step 1: Keywords by domain
    _task["progress"] = "正在抓取微博热搜..."
    domain_keywords = scrape_weibo_hot()
    flat_keywords = []
    for kws in domain_keywords.values():
        flat_keywords.extend(kws)
    kw_count = len(flat_keywords)

    # Step 2: Articles
    _task["progress"] = "正在检索公众号文章..."
    articles = search_articles_bulk()
    total_scraped = len(articles)

    # Step 3: Match keywords & assign domain
    _task["progress"] = f"正在匹配关键词...（{len(articles)}篇文章）"
    for art in articles:
        matched_list, primary_domain, primary_keyword = match_keywords_with_domain(art, domain_keywords)
        art["matched_keywords"] = [m["keyword"] for m in matched_list]
        art["matched_details"] = matched_list
        art["match_count"] = len(matched_list)
        art["domain"] = primary_domain
        art["primary_keyword"] = primary_keyword

    # Step 4: Score
    _task["progress"] = "AI 评分中..."
    for art in articles:
        art["ai_score"] = ai_score(art)
        art["time_weight"] = time_weight(art["pub_time"])
        art["total_score"] = round(
            art["match_count"] * 1.5 + art["ai_score"] + art["time_weight"], 2
        )

    # Step 5: Verify source authenticity
    verified = [art for art in articles if verify_article_source(art)]
    print(f"[Pipeline] Verified: {len(verified)} / {len(articles)} articles")

    # Step 6: Domain-balanced selection (每领域 ~10 篇，总计 50)
    _task["progress"] = "生成排行榜..."
    per_domain = 10
    domain_buckets = {d: [] for d in DOMAIN_NAMES}
    for art in verified:
        domain_buckets[art["domain"]].append(art)

    # 每个桶内按总分排序
    for d in DOMAIN_NAMES:
        domain_buckets[d].sort(key=lambda x: -x["total_score"])

    # 第一轮：每领域取 per_domain 篇
    top50 = []
    for d in DOMAIN_NAMES:
        top50.extend(domain_buckets[d][:per_domain])

    # 如果某领域不足 10 篇，用其他领域高分文章补齐
    if len(top50) < 50:
        used_titles = set(a["title"] for a in top50)
        remaining = [a for a in verified if a["title"] not in used_titles]
        remaining.sort(key=lambda x: -x["total_score"])
        for art in remaining:
            if len(top50) >= 50:
                break
            top50.append(art)

    # 最终按总分排序
    top50.sort(key=lambda x: -x["total_score"])
    top50 = top50[:50]

    # 清理内部字段
    for art in top50:
        art.pop("matched_details", None)
        art.pop("_source", None)

    # 统计各领域分布
    domain_dist = {d: 0 for d in DOMAIN_NAMES}
    for art in top50:
        domain_dist[art["domain"]] = domain_dist.get(art["domain"], 0) + 1

    return {
        "articles": top50,
        "stats": {
            "total_scraped": total_scraped,
            "keyword_count": kw_count,
            "top50_count": len(top50),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "accounts_covered": len(set(a["account"] for a in top50)),
            "domain_distribution": domain_dist,
        }
    }


def build_excel(data):
    """Generate Excel file in memory"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TOP50"

    hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cfont = Font(name="Arial", size=10)
    calign = Alignment(vertical="center", wrap_text=True)
    ccenter = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    gen_at = data["stats"]["generated_at"]
    t = ws.cell(row=1, column=1, value=f"公众号热点文章 TOP50 清单 | {gen_at}")
    t.font = Font(name="Arial", bold=True, size=14, color="1B2A4A")
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 35

    headers = ["排名", "领域", "公众号", "文章标题", "摘要", "匹配热搜", "发布时间", "匹配词数", "AI评分", "总分"]
    widths = [6, 10, 16, 42, 52, 20, 18, 12, 10, 10]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 28

    gold = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    alt1 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    alt2 = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    for idx, art in enumerate(data["articles"]):
        row = idx + 3
        rank = idx + 1
        fill = gold if rank <= 3 else (alt2 if rank % 2 == 0 else alt1)
        vals = [rank, art.get("domain", ""), art["account"], art["title"], art["summary"],
                art.get("primary_keyword", ""), art["pub_time"],
                art["match_count"], art["ai_score"], art["total_score"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = cfont; c.border = border; c.fill = fill
            c.alignment = ccenter if ci in [1, 2, 7, 8, 9, 10] else calign
        ws.row_dimensions[row].height = 42

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{len(data['articles'])+2}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Async Task State ───────────────────────────────────────
_task = {"running": False, "progress": "", "error": None}


def _run_generate():
    """Background thread: run the full pipeline"""
    try:
        _task["progress"] = "正在抓取微博热搜..."
        data = generate_top50()
        _cache["data"] = data
        _cache["time"] = time.time()
        _cache["excel"] = build_excel(data)
        _task["progress"] = "完成"
    except Exception as e:
        import traceback
        traceback.print_exc()
        _task["error"] = str(e)
        _task["progress"] = "出错"
    finally:
        _task["running"] = False


# ── Routes ─────────────────────────────────────────────────

@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/api/generate", methods=["POST"])
def api_generate():
    """Start generation in background thread, return immediately"""
    if _task["running"]:
        return jsonify({"status": "running", "progress": _task["progress"]})

    _task["running"] = True
    _task["error"] = None
    _task["progress"] = "启动中..."
    t = threading.Thread(target=_run_generate, daemon=True)
    t.start()
    return jsonify({"status": "started"})


@app.route("/api/status")
def api_status():
    """Poll this endpoint to check if generation is done"""
    if _task["running"]:
        return jsonify({"status": "running", "progress": _task["progress"]})

    if _task["error"]:
        return jsonify({"status": "error", "error": _task["error"],
                        "articles": [], "stats": {
                            "total_scraped": 0, "keyword_count": 0, "top50_count": 0,
                            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "accounts_covered": 0, "domain_distribution": {},
                        }})

    if _cache["data"]:
        return jsonify({"status": "done", **_cache["data"]})

    return jsonify({"status": "idle"})


@app.route("/api/download")
def api_download():
    if not _cache["data"]:
        data = generate_top50()
        _cache["data"] = data
        _cache["time"] = time.time()
        _cache["excel"] = build_excel(data)

    buf = build_excel(_cache["data"])
    ts = datetime.now().strftime("%Y%m%d")
    return send_file(buf, as_attachment=True,
                     download_name=f"top50_hot_articles_{ts}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)
